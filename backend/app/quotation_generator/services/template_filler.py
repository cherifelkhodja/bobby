"""Template filler service for Excel quotation documents."""

import io
import logging
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.quotation_generator.domain.entities import Quotation
from app.quotation_generator.domain.exceptions import TemplateFillerError

logger = logging.getLogger(__name__)


class TemplateFillerService:
    """Service for filling Excel templates with quotation data.

    This service uses openpyxl to fill Excel templates with quotation
    data. It supports Jinja2-like placeholder syntax: {{ variable_name }}.
    """

    # Regex pattern for placeholders: {{ variable_name }}
    PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*(\w+)\s*\}\}")

    def fill_template(
        self,
        template_content: bytes,
        quotation: Quotation,
        boond_reference: str | None = None,
    ) -> bytes:
        """Fill an Excel template with quotation data.

        Args:
            template_content: Template file content as bytes.
            quotation: Quotation entity with data.
            boond_reference: Optional BoondManager reference.

        Returns:
            Filled template as bytes.

        Raises:
            TemplateFillerError: If template filling fails.
        """
        try:
            # Load template from bytes
            workbook = load_workbook(io.BytesIO(template_content))

            # Get template context
            context = quotation.to_template_context(boond_reference)

            # Fill all worksheets
            for sheet in workbook.worksheets:
                self._fill_worksheet(sheet, context)

            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)

            logger.info(f"Filled template for quotation {quotation.resource_trigramme}")
            return output.getvalue()

        except Exception as e:
            logger.error(f"Template filling failed: {e}")
            raise TemplateFillerError(f"Failed to fill template: {str(e)}") from e

    def fill_template_file(
        self,
        template_path: Path,
        quotation: Quotation,
        output_path: Path,
        boond_reference: str | None = None,
    ) -> Path:
        """Fill template file and save to output path.

        Args:
            template_path: Path to template file.
            quotation: Quotation entity with data.
            output_path: Path for output file.
            boond_reference: Optional BoondManager reference.

        Returns:
            Path to filled template.

        Raises:
            TemplateFillerError: If filling fails.
            FileNotFoundError: If template not found.
        """
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")

        template_content = template_path.read_bytes()
        filled_content = self.fill_template(template_content, quotation, boond_reference)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(filled_content)

        return output_path

    def _fill_worksheet(self, sheet: Worksheet, context: dict[str, Any]) -> None:
        """Fill placeholders in a worksheet.

        Args:
            sheet: Worksheet to fill.
            context: Dictionary of variable values.
        """
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell.value = self._replace_placeholders(cell.value, context)

    def _replace_placeholders(self, text: str, context: dict[str, Any]) -> str:
        """Replace placeholders in text with context values.

        Args:
            text: Text containing {{ placeholder }} patterns.
            context: Dictionary of variable values.

        Returns:
            Text with placeholders replaced.
        """

        def replace_match(match: re.Match) -> str:
            variable_name = match.group(1)
            value = context.get(variable_name, "")
            return str(value) if value is not None else ""

        return self.PLACEHOLDER_PATTERN.sub(replace_match, text)

    def get_template_variables(self, template_content: bytes) -> list[str]:
        """Extract all variable names from a template.

        Args:
            template_content: Template file content as bytes.

        Returns:
            List of unique variable names found in template.
        """
        try:
            workbook = load_workbook(io.BytesIO(template_content))
            variables = set()

            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            matches = self.PLACEHOLDER_PATTERN.findall(cell.value)
                            variables.update(matches)

            return sorted(variables)

        except Exception as e:
            logger.error(f"Failed to extract template variables: {e}")
            raise TemplateFillerError(f"Failed to extract variables: {str(e)}") from e

    def validate_template(
        self,
        template_content: bytes,
        required_variables: list[str] | None = None,
    ) -> dict[str, Any]:
        """Validate a template and return analysis.

        Args:
            template_content: Template file content as bytes.
            required_variables: Optional list of required variable names.

        Returns:
            Dictionary with validation results:
            - is_valid: bool
            - variables_found: list of variable names
            - missing_variables: list of missing required variables
            - errors: list of error messages
        """
        result = {
            "is_valid": True,
            "variables_found": [],
            "missing_variables": [],
            "errors": [],
        }

        try:
            variables = self.get_template_variables(template_content)
            result["variables_found"] = variables

            if required_variables:
                missing = set(required_variables) - set(variables)
                if missing:
                    result["missing_variables"] = sorted(missing)
                    result["is_valid"] = False
                    result["errors"].append(f"Missing required variables: {', '.join(missing)}")

        except Exception as e:
            result["is_valid"] = False
            result["errors"].append(str(e))

        return result


# Default required variables for Thales PSTF template
THALES_PSTF_VARIABLES = [
    "thales_stakeholder",
    "procurement_buyer",
    "sow_reference",
    "object_of_need",
    "eacq_number",
    "reference",
    "sales_representative",
    "renewal",
    "initial_first_starting_date",
    "total_uo",
    "po_start_date",
    "po_end_date",
    "c22_domain",
    "c22_activity",
    "activity_country",
    "complexity_level",
    "gfa_max_price",
    "in_situ_ratio",
    "subcontracting",
    "tier2_supplier",
    "tier3_supplier",
    "additional_comments",
]
